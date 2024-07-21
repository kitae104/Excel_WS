import openpyxl
workbook = openpyxl.Workbook()        # 워크북 생성
ws = workbook.active               # 시트 생성 
ws.title = "python"                # 시트 타이틀 
ws["A1"] = "Name"                  # 특정 셀에 데이터 입력 
ws["C2"] = 20
ws["A3"] = "gil dong"
ws["F7"] = "test data"
workbook.save("input_test.xlsx")      # 파일 저장 