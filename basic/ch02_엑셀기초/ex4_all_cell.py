import openpyxl

file = "wss2.xlsx"
workbook = openpyxl.load_workbook(file)  # 엑셀 불러오기
ws = workbook.active                           # 첫번째 시트 활성화 하기

# 반복처리(모든 행과 열 접근)
for row in ws.iter_rows(values_only=True):
    # print(row)
    for value in row:
        if value is not None:
            print(value)

print("*" * 50)

# 스라이싱을 사용해서 부분 접근하기 
range = ws['A1': 'B2']  # 2 X 2
for row in range:
  for cell in row:
    print(cell.value)
