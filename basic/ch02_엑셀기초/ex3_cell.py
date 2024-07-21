import openpyxl

file = "wss.xlsx"
workbook = openpyxl.load_workbook("wss.xlsx")  # 엑셀 불러오기 
ws = workbook.active                           # 첫번째 시트 활성화 하기 

# 특정 셀에 접근하기 
ws.cell(row=5, column=1, value = "12345")      # 특정 위치에 데이터 넣기

ws['C1'] = 55
ws['D1'] = 10
ws['E1'] = "=C1 + D1"                          # 계산식

# 리스트 데이터 처리하기 
lst = [3,4,12,54,23,66,21,90,567,13]              

for i, n in enumerate(lst):
  ws.cell(row=i+1, column=2, value=n)

# 특정 열 가져오기 
column_values = [cell.value for cell in ws['B']]
print(column_values)

# 특정 행 가져오기 
row_values = [cell.value for cell in ws[1]]
print(row_values)

# 특정셀 접근하기 
v = ws["A1"].value
print(v)


workbook.save("wss2.xlsx")