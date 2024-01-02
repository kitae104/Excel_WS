import openpyxl
from openpyxl.styles import PatternFill, Font

wb = openpyxl.load_workbook("attendance.xlsx")  # 파일 불러오기
sheet = wb.active                         # 현재 활성화된 sheet 가져오기

for row in sheet.iter_rows(min_row = 1, max_row = 10, min_col =1, max_col=10): 
    for cell in row:
        if cell.value is not None and cell.value == "A":
            cell.fill = PatternFill(start_color="FF0055", end_color="FF0055", fill_type="solid")
            cell.font = Font(b=True)

wb.save("attendance.xlsx")