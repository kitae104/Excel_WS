from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

wb = load_workbook("file.xlsx")  # 파일 불러오기
ws = wb.active                         # 현재 활성화된 ws 가져오기

# 1~10행, 1~3열의 셀을 순회하면서 
for row in ws.iter_rows(min_row = 1, max_row = 10, min_col =1, max_col=3): 
    for cell in row:
        # 셀의 값이 40보다 크면 배경을 노란색으로, 글씨를 기울임꼴로 변경
        if cell.value is not None and cell.value > 40:
            cell.fill = PatternFill(start_color="ffccaa", end_color="123456", fill_type="solid")
            cell.font = Font(i=True)
        # 셀의 값이 20보다 작으면 배경을 빨간색으로, 글씨를 굵게 변경
        if cell.value is not None and cell.value < 20:
            cell.fill = PatternFill(start_color="FF0055", end_color="FF0055", fill_type="solid")
            cell.font = Font(b=True)
        # 셀의 값이 10보다 크고 20보다 작으면 배경을 노란색으로, 글씨를 굵고 기울임꼴로 변경
        if cell.value is not None and cell.value > 10 and cell.value < 20:
            cell.fill = PatternFill(start_color="ffccaa", end_color="123456", fill_type="solid")
            cell.font = Font(b= True, i = True, u = "single")

wb.save("file.xlsx")
