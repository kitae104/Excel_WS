import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("file.xlsx")
ws = wb.active

total_rows = ws.max_row
start_row = max(total_rows-3, 1)
end_row = total_rows

for row in range(start_row,end_row):
    for col in range(2,4): # A to D
        ws.cell(row=row, column = col).fill = PatternFill(start_color="ffccaa", end_color="123456", fill_type="solid")

wb.save("file.xlsx")