import openpyxl
from openpyxl.styles import PatternFill

workbook = openpyxl.load_workbook("show_fail.xlsx")
ws = workbook.active

for row in ws.iter_rows(min_row = 2, max_row = 10, min_col =4, max_col=9):
    for cell in row:
        if cell.value is not None and cell.value < 23:
            cell.fill = PatternFill(start_color="ffccaa", end_color="123456", fill_type="solid")
workbook.save("show fail.xlsx")