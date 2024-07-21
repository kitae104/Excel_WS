import openpyxl
import numpy 

wb = openpyxl.load_workbook("writeonlyfile.xlsx")
ws = wb.active

age_values = [cell.value for cell in ws['B']]
print(age_values[1:])
array_of_ages = numpy.array(age_values[1:])

sum = numpy.sum(array_of_ages)
print(sum)
mean = numpy.mean(array_of_ages)
print(mean)

ws['E1'].value = sum
ws['E2'].value = mean

wb.save("numpy.xlsx")