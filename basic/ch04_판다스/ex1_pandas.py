import openpyxl
import pandas 

wb = openpyxl.load_workbook("writeonlyfile.xlsx")
ws = wb.active

df = pandas.read_excel("writeonlyfile.xlsx", ws_name='ws')
Age = df['Age']
print(Age)

age_sum = Age.sum()
print(age_sum)

age_mean = Age.mean()
print(age_mean)

ws['D1'].value = age_sum
ws['D2'].value = age_mean

wb.save("pandas.xlsx")