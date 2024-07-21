import openpyxl 
from openpyxl.styles import Font

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "data"

# 병합하기 & 병합 풀기 
ws.merge_cells("B1:E1")
ws["B1"] = "엑셀 병합 테스트"
ws.unmerge_cells("B1:E1")

# 포맷 설정하기 
font1 = Font(
  name="Times New Roman",
  bold = True,
  italic=True,
    u = "double",
    color = "ff0011",
    size = 14,
    strike= True 
)

font2 = Font(
    name = "Arial",
    bold = False,
    italic=True,
    u = "single",
    color = "0011FF",
    size = 20,
    strike= False 
)

ws["A1"].font = font1
ws["B1"].font = font2

wb.save("merging.xlsx")