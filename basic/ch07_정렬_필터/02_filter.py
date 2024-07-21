# 필터
from openpyxl import load_workbook

wb = load_workbook("basic/ch07_정렬_필터/file.xlsx")
ws = wb.active

sp_column = f'B1:B{ws.max_row}' # A열의 데이터를 가져옴
ws.auto_filter.ref = sp_column  # 데이터의 범위를 지정
print(sp_column)
print(ws.dimensions)    # 데이터의 모든 범위를 출력
wb.save("basic/ch07_정렬_필터/filter_file.xlsx")