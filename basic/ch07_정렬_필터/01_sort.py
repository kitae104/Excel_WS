# 워크북의 데이터를 정렬(내림 차순)하여 다른 엑셀 파일에 저장하는 예제
from openpyxl import load_workbook

wb = load_workbook("basic/ch07_정렬_필터/file.xlsx")
ws = wb.active

data_list = list(ws.iter_rows(min_row=2, values_only=True)) # 2번째 줄부터 데이터를 가져옴
print(data_list)

sorted_data = sorted(data_list, key=lambda x: x[0], reverse=True) # 1번째 열을 기준으로 내림차순 정렬
print(sorted_data)

# ws.delete_rows(2, ws.max_row) # 2번째 줄부터 마지막 줄까지 삭제

for data in sorted_data:    # 정렬된 데이터를 삽입
    ws.append(data)         # append() 함수를 이용하여 데이터를 삽입

wb.save("basic/ch07_정렬_필터/sorted_file.xlsx")
