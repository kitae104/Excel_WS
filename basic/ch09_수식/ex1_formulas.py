from openpyxl import load_workbook

wb = load_workbook("basic/ch09_수식/data.xlsx")
ws = wb.active

ws['D1'].value = "=SUM(A1:A7)"       # 합계
ws['B1'].value = "=A1+A2"            # 덧셈
ws['B2'].value = (
    "=A1+A2"
)

var = 5
ws['B3'].value = f"=A1*{var}"        # f-string 사용 

# 함수 사용
ws['D2'].value = "=AVERAGE(A1:A7)"   # 평균
ws['D3'].value = "=MIN(A1:A7)"       # 최소값

# 조건식, 문자열 연결, 개수
ws['E1'].value = '=IF(A1>6, "True", "False")'    # 조건식
ws['F1'].value = '=CONCATENATE(A1," ",B1)'       # 문자열 연결
ws['C1'].value = '=COUNT(A1:A7)'     # 개수

wb.save("basic/ch09_수식/result_data.xlsx")