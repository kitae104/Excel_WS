# 엑셀 파일 로드하기 
import openpyxl
wb = openpyxl.load_workbook('test.xlsx')    # 엑셀 파일 로드 
ws = wb.active                           # 시트 활성화 
cell_value1 = ws['A1'].value             # 셀 값
cell_value2 = ws['B2'].value
print(cell_value1)
print(cell_value2)
