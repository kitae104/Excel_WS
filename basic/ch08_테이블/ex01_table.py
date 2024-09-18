import openpyxl 
from openpyxl.worksheet.table import Table
from openpyxl.styles import PatternFill

wb = openpyxl.Workbook()  # 새 워크북 생성
ws = wb.active # 현재 활성화된 sheet 가져오기
 
 # 데이터 준비 
data = [                     
  ["이름", "나이", "도시"],
  ["홍길동", 20, "서울"],
  ["김철수", 30, "인천"],
  ["박성진", 40, "대전"],
  ["이영희", 50, "대구"],
  ["최선이", 60, "부산"]
]

# 데이터 입력
for row in data:
  ws.append(row)

# 테이블 만들기 
table_range = f"A1:C{len(data)}" # A1 to C5
table = Table(displayName="MyTableData", ref=table_range) # 테이블 생성

# 행 삽입 후 데이터 입력 
ws.insert_rows(2)                           # 2번째 행에 삽입
ws.cell(row = 2 , column=1, value="New1") # 2번째 행 1번째 열에 데이터 입력
ws.cell(row = 2 , column=2, value="New2")       # 2번째 행 2번째 열에 데이터 입력
ws.cell(row = 2 , column=3, value="New3")   # 2번째 행 3번째 열에 데이터 입력

ws.insert_rows(4)
ws.cell(row = 4 , column=1, value="New4")
ws.cell(row = 4 , column=2, value="New5")
ws.cell(row = 4 , column=3, value="New6")

# 열 삽입 후 데이터 입력
ws.insert_cols(2)
ws.cell(row=1, column =2, value="테이블")

ws.cell(row = 2 , column=2, value=23)
ws.cell(row = 3 , column=2, value=24)
ws.cell(row = 4 , column=2, value=55)
ws.cell(row = 5 , column=2, value=13)
ws.cell(row = 6 , column=2, value=44)
ws.cell(row = 7 , column=2, value=65)

# 테이블 스타일 지정 - 2번째 행에 대해서만 적용
for cell in ws[2]:
    cell.fill = PatternFill(start_color="ffaabb", end_color="123456", fill_type="solid")

for cell in ws[7]:
    cell.fill = PatternFill(start_color="ffccbb", end_color="ff3456", fill_type="solid")

for row in ws.iter_rows(min_col = 3, max_col = 3):  # 3번째 열에 대해서만 적용
    for cell in row:
        cell.fill = PatternFill(start_color="ffaabb", end_color="123456", fill_type="darkTrellis")

ws.add_table(table) # 테이블 추가
wb.save("basic/ch08_테이블/table.xlsx") # 엑셀 파일 저장